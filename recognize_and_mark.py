import cv2
import numpy as np
import os
from datetime import datetime
from utils import load_model
import openpyxl
import time
from geopy.distance import geodesic
from geopy.geocoders import Nominatim
import geocoder

def get_user_location():
    """
    Fetch the user's current location using geocoder (IP-based).
    
    Returns:
    - tuple: (latitude, longitude) of the user's location.
    """
    try:
        g = geocoder.ip('me')  # Get location based on IP address
        if g.ok:
            print(g.latlng)
            return g.latlng
        else:
            print("Could not fetch location. Ensure you are connected to the internet.")
            return None
    except Exception as e:
        print(f"Failed to get user location: {e}")
        return None

def is_in_classroom(user_coords, classroom_coords, offset=50):
    """
    Check if the user is within the classroom range.

    Parameters:
    - user_coords (tuple): (latitude, longitude) of the user's location.
    - classroom_coords (tuple): (latitude, longitude) of the classroom.
    - offset (float): Offset range in meters.

    Returns:
    - bool: True if the user is within the classroom range, False otherwise.
    """
    distance = geodesic(user_coords, classroom_coords).meters
    print(f"Distance to classroom: {distance:.2f} meters")
    return distance <= offset

def recognize_face(model_path="models/", data_path="data/", excel_file="attendance.xlsx"):
    """
    Recognizes faces in real-time and marks attendance for recognized faces if within classroom location.

    Parameters:
    - model_path (str): Path to the trained model files.
    - data_path (str): Path to the face image folder.
    - excel_file (str): Path to the Excel file for attendance.
    """
    # Define classroom coordinates
    classroom_coords = (12.9719, 77.5936)

    # Load the trained model
    print("Loading model...")
    eigenfaces, mean_face, projections, labels = load_model(model_path)
    
    # Map label IDs to names
    label_names = {idx: name for idx, name in enumerate(os.listdir(data_path))}

    # Initialize webcam
    cap = cv2.VideoCapture(0)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    print("Press 'q' to quit.")

    recognized_faces = set()  # Keep track of already marked faces

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture image.")
            break

        # Convert to grayscale
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Detect faces
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50))

        for (x, y, w, h) in faces:
            # Crop and preprocess the face
            face = gray[y:y + h, x:x + w]
            face_resized = cv2.resize(face, (64, 64)).flatten()

            # Subtract the mean face
            face_centered = face_resized - mean_face

            # Project onto eigenfaces
            face_projection = np.dot(face_centered, eigenfaces)

            # Compare with stored projections
            distances = np.linalg.norm(projections - face_projection, axis=1)
            closest_idx = np.argmin(distances)
            recognized_label = labels[closest_idx]
            recognized_name = label_names.get(recognized_label, "Unknown")

            # Display the result
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(frame, recognized_name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)

            if recognized_name != "Unknown" and recognized_name not in recognized_faces:
                user_coords = get_user_location()
                if user_coords and is_in_classroom(user_coords, classroom_coords, offset=50):  # Offset in meters
                    # Mark as "present" for recognized faces within location range
                    print(f"Marking {recognized_name} as present.")
                    mark_attendance(recognized_name, status="present", excel_file=excel_file)
                    recognized_faces.add(recognized_name)
                else:
                    print(f"{recognized_name} is outside the classroom location.")

        # Show the frame
        cv2.imshow("Face Recognition", frame)

        # Exit on 'q' key
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

def mark_attendance(name, status, excel_file="attendance.xlsx"):
    """
    Marks attendance for the recognized person in an Excel file with status.

    Parameters:
    - name (str): Name of the recognized person.
    - status (str): Status of attendance ("present").
    - excel_file (str): Path to the Excel file.
    """
    # Check if the Excel file exists, if not, create a new one
    if not os.path.exists(excel_file):
        print(f"File {excel_file} does not exist. Creating a new file...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time", "Status"])  # Add headers
        workbook.save(excel_file)  # Save the new file

    # Open the existing Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Get the current date and time
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Check if the person is already marked present today
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == date:
            print(f"{name} is already marked present today.")
            return  # Person is already marked present

    # Append the attendance entry
    sheet.append([name, date, time, status])
    workbook.save(excel_file)
    print(f"Marked {status} attendance for {name} at {time}.")

if __name__ == "__main__":
    recognize_face()
