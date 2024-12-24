import cv2
import numpy as np
import os
from datetime import datetime
from utils import load_model
import openpyxl
import time

def recognize_face(model_path="models/", data_path="data/", excel_file="attendance.xlsx"):
    """
    Recognizes faces in real-time and marks attendance based on face duration.

    Parameters:
    - model_path (str): Path to the trained model files.
    - data_path (str): Path to the face image folder.
    - excel_file (str): Path to the Excel file for attendance.
    """
    # Load the trained model
    print("Loading model...")
    eigenfaces, mean_face, projections, labels = load_model(model_path)
    
    # Map label IDs to names
    label_names = {idx: name for idx, name in enumerate(os.listdir(data_path))}

    # Initialize webcam
    cap = cv2.VideoCapture(0)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    print("Press 'q' to quit.")

    face_tracker = {}  # Dictionary to track face ID and timestamps

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to capture image.")
            break

        # Convert to grayscale
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Detect faces
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50))

        for (x, y, w, h) in faces:
            # Crop and preprocess the face
            face = gray[y:y + h, x:x + w]
            face_resized = cv2.resize(face, (64, 64)).flatten()

            # Subtract the mean face
            face_centered = face_resized - mean_face

            # Project onto eigenfaces
            face_projection = np.dot(face_centered, eigenfaces)

            # Compare with stored projections
            distances = np.linalg.norm(projections - face_projection, axis=1)
            closest_idx = np.argmin(distances)
            recognized_label = labels[closest_idx]
            recognized_name = label_names.get(recognized_label, "Unknown")

            # Display the result
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(frame, recognized_name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)

            # Track the face duration
            current_time = time.time()

            if recognized_name != "Unknown":
                # Check if the person is already tracked in the frame
                if recognized_name not in face_tracker:
                    face_tracker[recognized_name] = {"start_time": current_time, "status": "partial"}

                # Calculate the duration face is present in the frame
                duration = current_time - face_tracker[recognized_name]["start_time"]

                if duration >= 10 and face_tracker[recognized_name]["status"] == "partial":
                    # Mark as "present" if the face stays for 10 seconds
                    print(f"{recognized_name} has been in the frame for {duration:.2f} seconds. Marking as present.")
                    mark_attendance(recognized_name, status="present", excel_file=excel_file)
                    face_tracker[recognized_name]["status"] = "present"

                elif duration < 10 and face_tracker[recognized_name]["status"] != "present":
                    # If the person leaves before 10 seconds, mark as "partial"
                    print(f"{recognized_name} has been in the frame for {duration:.2f} seconds. Marking as partial.")
                    mark_attendance(recognized_name, status="partial", excel_file=excel_file)

        # Show the frame
        cv2.imshow("Face Recognition", frame)

        # Exit on 'q' key
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

def mark_attendance(name, status, excel_file="attendance.xlsx"):
    """
    Marks attendance for the recognized person in an Excel file with status.

    Parameters:
    - name (str): Name of the recognized person.
    - status (str): Status of attendance ("present" or "partial").
    - excel_file (str): Path to the Excel file.
    """
    # Check if the Excel file exists, if not, create a new one
    if not os.path.exists(excel_file):
        print(f"File {excel_file} does not exist. Creating a new file...")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time", "Status"])  # Add headers
        workbook.save(excel_file)  # Save the new file

    # Open the existing Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Get the current date and time
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Check if the person is already marked present today
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == date:
            print(f"{name} is already marked present today.")
            return  # Person is already marked present

    # Append the attendance entry
    sheet.append([name, date, time, status])
    workbook.save(excel_file)
    print(f"Marked {status} attendance for {name} at {time}.")
